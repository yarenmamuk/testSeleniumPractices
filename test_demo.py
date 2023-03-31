from selenium import webdriver
from time import sleep 
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl

class Test_DemoClass:
    # her testten önce çağırılır
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get("https://www.saucedemo.com/")
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True) # klasör varsa kontrol et
        # günün tarihini al bu tarih ile bir klasör var mı kontrol et yoksa oluştur
    # her testen sonra çğırılır
    def teardown_method(self):
        self.driver.quit()

    def readData(self):
        print("x")
    # setup -> test_demoFunc -> teardown
    def test_demoFunc(self):
        # 3A Act Arrange Assert
        text = "Hello"
        assert text == "Hello"
    # setup -> test_demo2 -> teardown
    def test_demo2(self):
        assert True

    def getData(): #dekartörlerden çağırdığımız için içeriye self yazmıyoruz
        #veriyi al
        excelFile = openpyxl.load_workbook("data/invalid_login.xlsx")
        selectedSheet = excelFile["Sheet1"]

        totalRows = selectedSheet.max_row
        data = []
        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)

        return data

    @pytest.mark.parametrize("username,password", getData())
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name"))
        usernameInput = self.driver.find_element(By.ID, "user-name")
        self.waitForElementVisible((By.ID,"password"))
        passwordInput = self.driver.find_element(By.ID, "password")      
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID, "login-button")
        loginBtn.click()
        errorMessage = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        # magic string
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"


    def waitForElementVisible(self,locator):
        WebDriverWait(self.driver,10).until(ec.visibility_of_element_located(locator))